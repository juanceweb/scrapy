import os
import math
import ast
import pandas as pd
from openpyxl import load_workbook


def crear_diccionario_en_excel(instancia_diccionario, nombre_archivo):
    # Si el archivo no existe, guardar directamente el DataFrame en el archivo Excel

    df = pd.DataFrame.from_dict(instancia_diccionario)
    # Transponer el DataFrame para que cada diccionario sea una fila
    df = df.transpose()

    df = df.reindex(columns=list(instancia_diccionario["column_1"].keys()))
    # Guardar el DataFrame en un archivo Excel
    df.to_excel(nombre_archivo, index=False)


def obtener_ultima_fila_con_informacion(nombre_archivo, nombre_columna):
    # Cargar el archivo de Excel
    df = pd.read_excel(nombre_archivo)

    value = df[nombre_columna].notna().sum()

    return value


def guardar_diccionario_en_excel(instancia_diccionario, nombre_archivo):
    # Si el archivo ya existe, cargar el contenido existente en un DataFrame

    id = obtener_ultima_fila_con_informacion(nombre_archivo, "PropertyID")

    nuevo_id = id + 1

    instancia_diccionario["column_1"]["PropertyID"] = "TOOLS_" + str(nuevo_id).zfill(5)

    df_existente = pd.read_excel(nombre_archivo)

    df_nuevo = pd.DataFrame.from_dict(instancia_diccionario)

    df_nuevo = df_nuevo.transpose()

    df_nuevo = df_nuevo.reindex(columns=list(instancia_diccionario["column_1"].keys()))

    df_final = pd.concat([df_existente, df_nuevo])

    # Guardar el DataFrame actualizado en el archivo Excel
    df_final.to_excel(nombre_archivo, index=False)


def actualizar_fila_excel(fila, nueva_data, nombre_archivo):
    fila += 2
    try:
        # Cargar el archivo Excel existente
        book = load_workbook(nombre_archivo)

        # Seleccionar la hoja de trabajo
        hoja = book.active

        # Actualizar la fila con la nueva data
        for i, valor in enumerate(nueva_data.values(), start=1):
            if i != 1:
                if type(valor) == list:
                    valor = str(valor)
                hoja.cell(row=fila, column=i).value = valor

        # Guardar los cambios en el archivo Excel
        book.save(nombre_archivo)

    except FileNotFoundError:
        print("El archivo Excel no se encuentra.")
    except Exception as e:
        print("Error al procesar el archivo Excel:", e)


def verificar_existente_identico(
    name,
    path,
    cat_eng,
    desc_eng,
    prices_mod,
    price,
    key_eng,
    web,
    media_link,
    nombre_archivo,
):
    try:
        df = pd.read_excel(nombre_archivo)

        lista_tools = df["Name"].tolist()

        try:
            indice = lista_tools.index(name)
            fila_especifica = df.iloc[indice]

            cambios = 0

            if set(ast.literal_eval(fila_especifica["Paths"])) != set(path):
                cambios = 1

            elif set(ast.literal_eval(fila_especifica["CategoriesENG"])) != set(
                cat_eng
            ):
                cambios = 1

            elif fila_especifica["Name"] != name:
                cambios = 1

            elif fila_especifica["DescriptionEnglish"] != desc_eng:
                cambios = 1

            elif fila_especifica["Prices"] != prices_mod:
                if math.isnan(fila_especifica["Prices"]) and prices_mod == "":
                    pass
                else:
                    cambios = 1

            elif set(ast.literal_eval(fila_especifica["BusinessModel"])) != set(price):
                cambios = 1

            elif set(ast.literal_eval(fila_especifica["KeywordsENG"])) != set(key_eng):
                cambios = 1

            elif fila_especifica["Link"] != web:
                cambios = 1

            elif fila_especifica["Media"] != media_link:
                cambios = 1

            if cambios == 0:
                return "EXISTE SIN CAMBIOS", False
            else:
                return "CAMBIOS", indice

        except Exception as e:
            return "NO EXISTE", False

    except FileNotFoundError:
        print("El archivo Excel no se encuentra.")
    except Exception as e:
        print("Error al procesar el archivo Excel:", e)


def verificar_existencia_excel(nombre_archivo):
    return os.path.isfile(nombre_archivo)
